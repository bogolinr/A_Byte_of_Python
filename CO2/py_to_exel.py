from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

wb = Workbook()
wb.create_sheet(title='Моделирование', index = 0)
sheet = wb['Моделирование']

sheet['A1'] = 'температура'
sheet['B1'] = 'время'
for i in range(1,11):
	cell = sheet.cell(row = i+1, column=1)
	cell.value = i *i
	cell = sheet.cell(row = i+1, column=2)
	cell.value = i

# chart = BarChart()
# chart.title='name'
# data = Reference(sheet, min_col=1,min_row=1,max_col=1, max_row=11)

# chart.add_data(data, titles_from_data=True)
# sheet.add_chart(chart, 'C2')
wb.save('example.xlsx')